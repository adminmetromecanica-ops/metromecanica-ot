import os
import re
import subprocess
import tempfile
import datetime
import shutil
import zipfile
from flask import Blueprint, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

certbot_bp = Blueprint('certbot', __name__)


def fmt_val(val, number_format=None):
    if val is None or isinstance(val, bool) or isinstance(val, str):
        return val
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        if not number_format or number_format in ("General", "@"):
            if isinstance(val, float) and val == int(val):
                return int(val)
            return val
        m = re.search(r'\.([0#]+)', number_format)
        decimales = len(m.group(1)) if m else 0
        redondeado = round(float(val), decimales)
        if decimales == 0:
            return int(redondeado)
        s = f"{redondeado:.{decimales}f}"
        return s.replace(".", ",")
    return val


def leer_certificado(ruta_excel):
    wb_vals = load_workbook(ruta_excel, read_only=False, data_only=True)
    cert_name = next(
        (s for s in wb_vals.sheetnames if s.upper() == "CERTIFICADO"),
        wb_vals.sheetnames[-1]
    )
    ws_vals = wb_vals[cert_name]
    valores = {}
    for row in ws_vals.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                valores[cell.coordinate] = (cell.value, cell.number_format)
    wb_vals.close()
    resultado = {}
    for coord, (val, fmt) in valores.items():
        resultado[coord] = fmt_val(val, fmt)
    return resultado, cert_name


def preparar_para_pdf(ruta_excel, tmpdir):
    valores_cert, cert_name = leer_certificado(ruta_excel)

    # Paso 1: copiar original preservando charts/drawings
    ruta_copia = os.path.join(tmpdir, "trabajo.xlsm")
    shutil.copy2(ruta_excel, ruta_copia)

    # Paso 2: inyectar valores estáticos con openpyxl
    wb = load_workbook(ruta_copia, data_only=False, keep_vba=True)
    ws = wb[cert_name]

    for coord, val in valores_cert.items():
        try:
            cell = ws[coord]
            if isinstance(cell, MergedCell):
                for rng in ws.merged_cells.ranges:
                    if coord in rng:
                        master = wb[cert_name].cell(row=rng.min_row, column=rng.min_col)
                        master.value = val
                        break
            else:
                cell.value = val
        except Exception:
            pass

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            try:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = None
            except Exception:
                pass

    # Obtener índice de hoja CERTIFICADO en workbook.xml
    cert_idx = wb.sheetnames.index(cert_name)

    wb.save(ruta_copia)
    wb.close()

    # Paso 3: manipular el ZIP para dejar SOLO la hoja CERTIFICADO
    # Mapear sheet names → sheet xml files
    ruta_out = os.path.join(tmpdir, "certificado_final.xlsx")

    with zipfile.ZipFile(ruta_copia, 'r') as zin:
        all_files = zin.namelist()

        # Leer workbook.xml para obtener relaciones de hojas
        wb_xml = zin.read('xl/workbook.xml').decode('utf-8')
        wb_rels = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')

        # Identificar sheet files de hojas que NO son CERTIFICADO
        import xml.etree.ElementTree as ET
        ns = {'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
              'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

        tree_wb  = ET.fromstring(wb_xml)
        tree_rel = ET.fromstring(wb_rels)

        # Construir mapa rId → target
        rid_to_target = {}
        for rel in tree_rel.findall('r:Relationship', ns):
            rid_to_target[rel.get('Id')] = rel.get('Target')

        # Identificar rIds de hojas que NO son CERTIFICADO
        sheets_to_remove = set()
        sheet_rids_to_remove = set()
        for sheet in tree_wb.findall('.//x:sheet', ns):
            name = sheet.get('name')
            rid  = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            if name.upper() != cert_name.upper():
                target = rid_to_target.get(rid, '')
                sheets_to_remove.add(target.replace('../', 'xl/').replace('xl/xl/', 'xl/'))
                sheet_rids_to_remove.add(rid)

        # Archivos a excluir del ZIP final
        exclude = set()
        for f in all_files:
            for s in sheets_to_remove:
                sheet_file = s if s.startswith('xl/') else f'xl/{s}'
                if f == sheet_file or f == sheet_file.replace('xl/worksheets/', 'xl/worksheets/'):
                    exclude.add(f)
                # También excluir _rels del sheet
                sheet_base = os.path.basename(sheet_file)
                if f == f'xl/worksheets/_rels/{sheet_base}.rels':
                    exclude.add(f)

        # Escribir nuevo ZIP solo con hoja CERTIFICADO
        with zipfile.ZipFile(ruta_out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in all_files:
                if item in exclude:
                    continue
                data = zin.read(item)

                # En workbook.xml: eliminar referencias a hojas removidas
                if item == 'xl/workbook.xml':
                    root = ET.fromstring(data.decode('utf-8'))
                    sheets_elem = root.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets')
                    if sheets_elem is not None:
                        for sheet in list(sheets_elem):
                            rid = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                            if rid in sheet_rids_to_remove:
                                sheets_elem.remove(sheet)
                    data = ET.tostring(root, encoding='utf-8', xml_declaration=True)

                # En workbook.xml.rels: eliminar relaciones de hojas removidas
                if item == 'xl/_rels/workbook.xml.rels':
                    root = ET.fromstring(data.decode('utf-8'))
                    for rel in list(root):
                        if rel.get('Id') in sheet_rids_to_remove:
                            root.remove(rel)
                    data = ET.tostring(root, encoding='utf-8', xml_declaration=True)

                zout.writestr(item, data)

    return ruta_out


def construir_nombre(ruta_excel, nombre_archivo):
    n_cert = magnitud = equipo = cliente = ot = ""
    try:
        wb = load_workbook(ruta_excel, read_only=True, data_only=True)
        if "CALIBRACION" in wb.sheetnames:
            cal = wb["CALIBRACION"]
            def g(coord):
                v = cal[coord].value
                return str(v).strip() if v else ""
            n_cert   = g("B150")
            magnitud = g("B151")
            equipo   = g("B152")
            cliente  = g("B153")
            ot       = g("B154")
        wb.close()
    except Exception:
        pass

    if not n_cert:
        for parte in nombre_archivo.upper().replace(".XLSM","").replace(".XLSX","").split("_"):
            if len(parte) > 5 and '-' in parte and any(
                parte.startswith(p) for p in ['MLL','MLF','MLE','MLP','MLT','MLM','MLQ','MLC','MLB']
            ):
                n_cert = parte
                break

    fecha  = datetime.date.today().strftime("%Y%m%d")
    nombre = f"{n_cert}_{magnitud}_{equipo}_{cliente}_{ot}_{fecha}.pdf"
    for c in ['\\','/',':', '*','?','"','<','>','|',' ','\n','\r']:
        nombre = nombre.replace(c, '_')
    while '__' in nombre:
        nombre = nombre.replace('__', '_')
    return nombre[:180]


@certbot_bp.route('/generar-certificado', methods=['POST'])
def generar_certificado():
    if 'file' not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400

    archivo = request.files['file']
    nombre  = archivo.filename

    with tempfile.TemporaryDirectory() as tmpdir:
        ruta_excel = os.path.join(tmpdir, nombre)
        archivo.save(ruta_excel)

        try:
            ruta_xlsx = preparar_para_pdf(ruta_excel, tmpdir)
        except Exception as e:
            return jsonify({"error": f"Error preparando archivo: {str(e)}"}), 500

        env = os.environ.copy()
        env["LANG"]       = "es_PE.UTF-8"
        env["LC_ALL"]     = "es_PE.UTF-8"
        env["LC_NUMERIC"] = "es_PE.UTF-8"

        cmd = [
            "libreoffice", "--headless",
            "--convert-to", "pdf",
            "--outdir", tmpdir,
            ruta_xlsx
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=90, env=env)
        if result.returncode != 0:
            return jsonify({"error": "Error LibreOffice", "detalle": result.stderr}), 500

        pdf_generado = os.path.join(tmpdir, "certificado_final.pdf")
        if not os.path.exists(pdf_generado):
            return jsonify({"error": "PDF no generado"}), 500

        pdf_final = os.path.join(tmpdir, construir_nombre(ruta_excel, nombre))

        try:
            from pypdf import PdfReader, PdfWriter
            reader = PdfReader(pdf_generado)
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            with open(pdf_final, "wb") as f:
                writer.write(f)
        except Exception:
            pdf_final = pdf_generado

        return send_file(
            pdf_final,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=os.path.basename(pdf_final)
        )
